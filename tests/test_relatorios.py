import io
from datetime import date

import openpyxl
import pandas as pd
import pytest

from cronograma.curva_s import gerar_curva_s
from cronograma.metricas import calcular_indicadores, gerar_percepcoes
from cronograma.portfolio import (
    calcular_indicadores_portfolio,
    gerar_curva_s_portfolio,
    indicadores_consolidados,
    tabela_comparativa,
)
from cronograma.relatorios import (
    gerar_excel,
    gerar_excel_portfolio,
    gerar_excel_tabela,
    gerar_pdf,
    gerar_pdf_portfolio,
    tabela_tarefas,
)

DATA_STATUS = date(2026, 7, 24)


@pytest.fixture()
def indicadores(projeto_com_custo):
    return calcular_indicadores(projeto_com_custo, DATA_STATUS)


@pytest.fixture()
def curva(projeto_com_custo, indicadores):
    return gerar_curva_s(projeto_com_custo, DATA_STATUS, percentual_concluido_alvo=indicadores.percentual_concluido)


def _abrir(dados: bytes):
    return openpyxl.load_workbook(io.BytesIO(dados))


def test_excel_completo_abre_com_planilhas(projeto_com_custo, indicadores, curva):
    wb = _abrir(gerar_excel(projeto_com_custo, indicadores, curva))
    assert wb.sheetnames == ["Resumo", "Tarefas", "Recursos", "Curva S"]


def test_excel_formatacao_cabecalho_e_congelamento(projeto_com_custo, indicadores, curva):
    wb = _abrir(gerar_excel(projeto_com_custo, indicadores, curva))
    ws = wb["Tarefas"]
    assert ws.freeze_panes == "A2"
    assert ws["A1"].font.bold
    assert ws["A1"].fill.fgColor.rgb.endswith("2E4053")


def test_excel_formatos_numericos(projeto_com_custo, indicadores, curva):
    wb = _abrir(gerar_excel(projeto_com_custo, indicadores, curva))
    ws = wb["Tarefas"]
    cabecalho = [c.value for c in ws[1]]
    col_data = cabecalho.index("Início") + 1
    col_custo = cabecalho.index("Custo") + 1
    col_pct = cabecalho.index("% Concluído") + 1
    assert ws.cell(row=2, column=col_data).number_format == "DD/MM/YYYY"
    assert "R$" in ws.cell(row=2, column=col_custo).number_format
    assert "%" in ws.cell(row=2, column=col_pct).number_format


def test_excel_destaca_linhas_criticas_atrasadas(projeto_com_custo, indicadores, curva):
    wb = _abrir(gerar_excel(projeto_com_custo, indicadores, curva))
    ws = wb["Tarefas"]
    cabecalho = [c.value for c in ws[1]]
    col_critica = cabecalho.index("Crítica") + 1
    col_atrasada = cabecalho.index("Atrasada") + 1
    for linha in range(2, ws.max_row + 1):
        critica = ws.cell(row=linha, column=col_critica).value == "Sim"
        atrasada = ws.cell(row=linha, column=col_atrasada).value == "Sim"
        cor = ws.cell(row=linha, column=1).fill.fgColor.rgb
        if atrasada and critica:
            assert cor.endswith("F1948A")
        elif atrasada:
            assert cor.endswith("F8D7DA")
        elif critica:
            assert cor.endswith("D7BDE2")


def test_excel_tabela_status_colorido():
    df = pd.DataFrame(
        [
            {"Tarefa": "A", "Status": "Concluída"},
            {"Tarefa": "B", "Status": "Crítica atrasada"},
        ]
    )
    ws = _abrir(gerar_excel_tabela(df, "Gantt"))["Gantt"]
    assert ws.cell(row=2, column=1).fill.fgColor.rgb.endswith("D4EFDF")
    assert ws.cell(row=3, column=1).fill.fgColor.rgb.endswith("F1948A")


def test_excel_tabela_vazia_nao_quebra():
    dados = gerar_excel_tabela(pd.DataFrame(), "Vazio")
    assert _abrir(dados).sheetnames == ["Vazio"]


def test_excel_tabela_tarefas_reais(projeto_com_custo):
    df = tabela_tarefas(projeto_com_custo)
    ws = _abrir(gerar_excel_tabela(df, "Tarefas"))["Tarefas"]
    assert ws.max_row == len(df) + 1


def test_pdf_gera_bytes(projeto_com_custo, indicadores, curva):
    percepcoes = [texto for _, texto in gerar_percepcoes(projeto_com_custo, indicadores, idioma="pt")]
    dados = gerar_pdf(projeto_com_custo, indicadores, percepcoes, curva)
    assert dados.startswith(b"%PDF")


def test_excel_e_pdf_portfolio(projeto_com_custo, projeto_sem_custo):
    projetos = {"a.xml": projeto_com_custo, "b.xml": projeto_sem_custo}
    ind = calcular_indicadores_portfolio(projetos)
    tabela = tabela_comparativa(projetos, ind, idioma="pt")
    consolidado = indicadores_consolidados(projetos, ind)
    curva_port = gerar_curva_s_portfolio(projetos, ind)

    wb = _abrir(gerar_excel_portfolio(tabela, consolidado, curva_port))
    assert "Resumo Portfólio" in wb.sheetnames
    assert "Comparativo" in wb.sheetnames

    pdf = gerar_pdf_portfolio(tabela, consolidado, curva_port)
    assert pdf.startswith(b"%PDF")
