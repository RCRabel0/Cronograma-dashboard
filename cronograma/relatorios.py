import io
from datetime import date, datetime

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .metricas import Indicadores, dias_atraso_tarefa, formatar_valor
from .modelos import Projeto, Tarefa

CORES_STATUS = {
    "Concluída": "#2E8B57",
    "No prazo": "#5B8DB8",
    "Crítica": "#4A235A",
    "Atrasada": "#E67E22",
    "Crítica atrasada": "#C0392B",
}


def status_tarefa(t: Tarefa) -> str:
    if t.percentual_concluido >= 100:
        return "Concluída"
    if t.atrasada and t.critica:
        return "Crítica atrasada"
    if t.atrasada:
        return "Atrasada"
    if t.critica:
        return "Crítica"
    return "No prazo"


_COR_CABECALHO = "2E4053"
_COR_FAIXA = "F2F3F4"
_COR_BORDA = "BFBFBF"
_COR_MARCO = "F9E79F"

# Mesmas cores usadas para destacar linhas na aba Tarefas do app (crítica+atrasada,
# atrasada, crítica), para que o Excel exportado bata visualmente com a tela.
_CORES_STATUS_TAREFA = {
    (True, True): "F1948A",
    (True, False): "F8D7DA",
    (False, True): "D7BDE2",
}

# Cores para a coluna "Status" do relatório do Gantt (valores em PT ou EN, conforme
# o idioma da interface no momento da exportação).
_CORES_STATUS_TEXTO = {
    "Concluída": "D4EFDF", "Completed": "D4EFDF",
    "No prazo": "D6EAF8", "On track": "D6EAF8",
    "Crítica": "D7BDE2", "Critical": "D7BDE2",
    "Atrasada": "F8D7DA", "Delayed": "F8D7DA",
    "Crítica atrasada": "F1948A", "Critical delayed": "F1948A",
}

# Cores para a coluna "Status" do checklist de qualidade.
_CORES_CONFORMIDADE = {
    "Conforme": "D4EFDF",
    "Parcial": "FCF3CF",
    "Não Conforme": "F8D7DA",
}


def _cor_linha(linha: dict) -> str | None:
    """Decide a cor de destaque de uma linha, reaproveitando a mesma lógica (e as
    mesmas cores) já usadas na aba Tarefas do app, ou a coluna 'Status' quando presente
    (relatório do Gantt ou checklist de qualidade)."""
    if "Atrasada" in linha and "Crítica" in linha:
        atrasada = linha.get("Atrasada") == "Sim"
        critica = linha.get("Crítica") == "Sim"
        cor = _CORES_STATUS_TAREFA.get((atrasada, critica))
        if cor:
            return cor
        if linha.get("Marco") == "Sim":
            return _COR_MARCO
        return None
    status = linha.get("Status")
    if status in _CORES_STATUS_TEXTO:
        return _CORES_STATUS_TEXTO[status]
    if status in _CORES_CONFORMIDADE:
        return _CORES_CONFORMIDADE[status]
    return None


def _formatar_planilha(ws: Worksheet, df: pd.DataFrame) -> None:
    """Aplica uma formatação consistente a uma planilha recém-escrita via
    DataFrame.to_excel: cabeçalho destacado, largura de coluna ajustada ao conteúdo,
    congelamento do cabeçalho, bordas, faixas de cor alternadas e formatos numéricos
    (datas, percentual, moeda) — além de destacar linhas de tarefas críticas/atrasadas
    ou por status, replicando as cores já usadas na interface.
    """
    if df.empty:
        return

    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    preenchimento_cabecalho = PatternFill("solid", fgColor=_COR_CABECALHO)
    lado = Side(style="thin", color=_COR_BORDA)
    borda_fina = Border(left=lado, right=lado, top=lado, bottom=lado)
    alinhamento_cabecalho = Alignment(horizontal="center", vertical="center", wrap_text=True)
    n_linhas = len(df)

    for col_idx, nome_coluna in enumerate(df.columns, start=1):
        celula_cabecalho = ws.cell(row=1, column=col_idx)
        celula_cabecalho.font = fonte_cabecalho
        celula_cabecalho.fill = preenchimento_cabecalho
        celula_cabecalho.alignment = alinhamento_cabecalho
        celula_cabecalho.border = borda_fina

        valores = df[nome_coluna]
        amostra = valores.dropna()
        eh_data = not amostra.empty and isinstance(amostra.iloc[0], (date, datetime))
        eh_percentual = "%" in str(nome_coluna)
        eh_moeda = "custo" in str(nome_coluna).lower()

        maior_texto = len(str(nome_coluna))
        for valor in valores:
            texto = "" if pd.isna(valor) else str(valor)
            maior_texto = max(maior_texto, len(texto))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(maior_texto + 2, 10), 50)

        if eh_data:
            numero_formato = "DD/MM/YYYY"
        elif eh_percentual:
            numero_formato = '0.0"%"'
        elif eh_moeda:
            numero_formato = '"R$" #,##0.00'
        elif pd.api.types.is_float_dtype(valores):
            numero_formato = "#,##0.00"
        else:
            numero_formato = None

        if numero_formato:
            for linha_idx in range(2, n_linhas + 2):
                ws.cell(row=linha_idx, column=col_idx).number_format = numero_formato

    n_colunas = len(df.columns)
    for pos, (_, linha) in enumerate(df.iterrows()):
        linha_idx = pos + 2
        cor_status = _cor_linha(linha.to_dict())
        cor_fundo = cor_status or (_COR_FAIXA if pos % 2 == 1 else None)
        for col_idx in range(1, n_colunas + 1):
            celula = ws.cell(row=linha_idx, column=col_idx)
            celula.border = borda_fina
            if cor_fundo:
                celula.fill = PatternFill("solid", fgColor=cor_fundo)

    ws.freeze_panes = "A2"


def gerar_excel_tabela(df: pd.DataFrame, nome_planilha: str = "Tarefas") -> bytes:
    """Gera um Excel simples de uma única planilha a partir de um DataFrame já pronto para exibição."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=nome_planilha[:31], index=False)
        _formatar_planilha(writer.sheets[nome_planilha[:31]], df)
    return buffer.getvalue()


def tabela_tarefas(projeto: Projeto) -> pd.DataFrame:
    linhas = []
    for t in projeto.tarefas_detalhe:
        linhas.append(
            {
                "ID": t.id,
                "Tarefa": t.nome,
                "Início": t.inicio,
                "Término": t.termino,
                "Início (linha de base)": t.inicio_linha_base,
                "Término (linha de base)": t.termino_linha_base,
                "% Concluído": t.percentual_concluido,
                "Marco": "Sim" if t.marco else "Não",
                "Crítica": "Sim" if t.critica else "Não",
                "Atrasada": "Sim" if t.atrasada else "Não",
                "Duração Planejada (h)": t.duracao_linha_base_horas or t.duracao_horas,
                "Duração Real (h)": t.duracao_real_horas,
                "Custo": t.custo,
                "Custo Real": t.custo_real,
                "Recursos": ", ".join(t.recursos),
            }
        )
    return pd.DataFrame(linhas)


def tabela_recursos(projeto: Projeto) -> pd.DataFrame:
    linhas = [
        {"Recurso": r.nome, "Trabalho (horas)": r.trabalho_horas, "Custo": r.custo}
        for r in projeto.recursos
    ]
    return pd.DataFrame(linhas)


def gerar_excel(projeto: Projeto, indicadores: Indicadores, curva_s: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()

    rotulo_ac = "Custo Real (AC)" if indicadores.unidade == "R$" else "Trabalho Real (AC)"
    rotulo_eac = "Custo Previsto ao Término (EAC)" if indicadores.unidade == "R$" else "Trabalho Previsto ao Término (EAC)"
    rotulo_cpi = "CPI (índice de custo)" if indicadores.unidade == "R$" else "CPI (índice de eficiência)"

    resumo = pd.DataFrame(
        {
            "Indicador": [
                "Projeto",
                "SPI (índice de prazo)",
                rotulo_cpi,
                "% Concluído",
                "Valor Planejado (PV)",
                "Valor Agregado (EV)",
                rotulo_ac,
                "SV (variação de prazo, EV - PV)",
                "CV (variação de custo, EV - AC)",
                "Atraso (dias)",
                rotulo_eac,
                "Tarefas Atrasadas",
                "Tarefas Críticas Atrasadas",
                "Total de Tarefas",
            ],
            "Valor": [
                projeto.nome,
                f"{indicadores.spi:.2f}" if indicadores.spi is not None else "N/D",
                f"{indicadores.cpi:.2f}" if indicadores.cpi is not None else "N/D",
                f"{indicadores.percentual_concluido:.1f}%",
                formatar_valor(indicadores.pv_total, indicadores.unidade),
                formatar_valor(indicadores.ev_total, indicadores.unidade),
                formatar_valor(indicadores.ac_total, indicadores.unidade),
                formatar_valor(indicadores.variacao_prazo, indicadores.unidade),
                formatar_valor(indicadores.variacao_custo, indicadores.unidade),
                indicadores.atraso_dias,
                formatar_valor(indicadores.custo_previsto_total, indicadores.unidade) if indicadores.custo_previsto_total else "N/D",
                indicadores.tarefas_atrasadas,
                indicadores.tarefas_criticas_atrasadas,
                indicadores.total_tarefas,
            ],
        }
    )

    colunas_grafico = ["Linha de Base", "Realizado / Previsto"]
    outras_colunas = [c for c in curva_s.columns if c not in colunas_grafico]
    curva_s_ordenada = curva_s[colunas_grafico + outras_colunas]

    df_tarefas = tabela_tarefas(projeto)
    df_recursos = tabela_recursos(projeto)
    df_curva = curva_s_ordenada.reset_index()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        resumo.to_excel(writer, sheet_name="Resumo", index=False)
        df_tarefas.to_excel(writer, sheet_name="Tarefas", index=False)
        df_recursos.to_excel(writer, sheet_name="Recursos", index=False)
        df_curva.to_excel(writer, sheet_name="Curva S", index=False)

        _formatar_planilha(writer.sheets["Resumo"], resumo)
        _formatar_planilha(writer.sheets["Tarefas"], df_tarefas)
        _formatar_planilha(writer.sheets["Recursos"], df_recursos)
        _formatar_planilha(writer.sheets["Curva S"], df_curva)

        planilha_curva = writer.sheets["Curva S"]
        grafico = LineChart()
        grafico.title = "Curva S - Linha de Base x Realizado/Previsto"
        grafico.y_axis.title = "Custo acumulado" if indicadores.unidade == "R$" else "Trabalho acumulado (horas)"
        grafico.x_axis.title = "Data"
        n_linhas = len(curva_s_ordenada) + 1
        dados = Reference(planilha_curva, min_col=2, max_col=1 + len(colunas_grafico), min_row=1, max_row=n_linhas)
        categorias = Reference(planilha_curva, min_col=1, min_row=2, max_row=n_linhas)
        grafico.add_data(dados, titles_from_data=True)
        grafico.set_categories(categorias)
        planilha_curva.add_chart(grafico, "G2")

    return buffer.getvalue()


def gerar_excel_portfolio(tabela_comparativa: pd.DataFrame, consolidado: dict, curva_portfolio: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()

    resumo = pd.DataFrame(
        {
            "Indicador": [
                "Total de Projetos",
                "Projetos Atrasados",
                "% Concluído (consolidado)",
                "SPI (consolidado)",
                "CPI (consolidado)",
                "Tarefas Atrasadas",
                "Tarefas Críticas Atrasadas",
                "Total de Tarefas",
            ],
            "Valor": [
                consolidado["total_projetos"],
                consolidado["projetos_atrasados"],
                f"{consolidado['percentual_concluido']:.1f}%",
                f"{consolidado['spi']:.2f}" if consolidado["spi"] is not None else "N/D",
                f"{consolidado['cpi']:.2f}" if consolidado["cpi"] is not None else "N/D",
                consolidado["tarefas_atrasadas"],
                consolidado["tarefas_criticas_atrasadas"],
                consolidado["total_tarefas"],
            ],
        }
    )

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        resumo.to_excel(writer, sheet_name="Resumo Portfólio", index=False)
        tabela_comparativa.to_excel(writer, sheet_name="Comparativo", index=False)
        _formatar_planilha(writer.sheets["Resumo Portfólio"], resumo)
        _formatar_planilha(writer.sheets["Comparativo"], tabela_comparativa)

        if not curva_portfolio.empty:
            df_curva_portfolio = curva_portfolio.reset_index(names="Data")
            df_curva_portfolio.to_excel(writer, sheet_name="Curva S Portfólio", index=False)
            _formatar_planilha(writer.sheets["Curva S Portfólio"], df_curva_portfolio)

            planilha_curva = writer.sheets["Curva S Portfólio"]
            grafico = LineChart()
            grafico.title = "Curva S Consolidada do Portfólio (%)"
            grafico.y_axis.title = "% Concluído (acumulado)"
            grafico.x_axis.title = "Data"
            n_linhas = len(curva_portfolio) + 1
            dados = Reference(planilha_curva, min_col=2, max_col=3, min_row=1, max_row=n_linhas)
            categorias = Reference(planilha_curva, min_col=1, min_row=2, max_row=n_linhas)
            grafico.add_data(dados, titles_from_data=True)
            grafico.set_categories(categorias)
            planilha_curva.add_chart(grafico, "F2")

    return buffer.getvalue()


def _grafico_curva_s_portfolio_png(curva_portfolio: pd.DataFrame) -> bytes:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(16, 8))
    for coluna in ["Linha de Base", "Realizado / Previsto"]:
        estilo = "--" if coluna == "Linha de Base" else "-"
        ax.plot(curva_portfolio.index, curva_portfolio[coluna], label=coluna, linewidth=2, linestyle=estilo)
    ax.set_title("Curva S Consolidada do Portfólio")
    ax.set_xlabel("Data")
    ax.set_ylabel("% Concluído (acumulado)")
    ax.legend()
    ax.grid(True, alpha=0.3)
    fig.autofmt_xdate()

    saida = io.BytesIO()
    fig.savefig(saida, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    return saida.getvalue()


def gerar_pdf_portfolio(tabela_comparativa: pd.DataFrame, consolidado: dict, curva_portfolio: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=1.2 * cm, bottomMargin=1.2 * cm, leftMargin=1.2 * cm, rightMargin=1.2 * cm,
    )
    estilos = getSampleStyleSheet()
    elementos = []
    largura_util = landscape(A4)[0] - 2.4 * cm

    elementos.append(Paragraph("Relatório de Portfólio de Projetos", estilos["Title"]))
    elementos.append(Spacer(1, 0.5 * cm))

    dados_indicadores = [
        ["Indicador", "Valor"],
        ["Total de Projetos", str(consolidado["total_projetos"])],
        ["Projetos Atrasados", str(consolidado["projetos_atrasados"])],
        ["% Concluído (consolidado)", f"{consolidado['percentual_concluido']:.1f}%"],
        ["SPI (consolidado)", f"{consolidado['spi']:.2f}" if consolidado["spi"] is not None else "N/D"],
        ["CPI (consolidado)", f"{consolidado['cpi']:.2f}" if consolidado["cpi"] is not None else "N/D"],
        ["Tarefas Atrasadas", str(consolidado["tarefas_atrasadas"])],
        ["Tarefas Críticas Atrasadas", str(consolidado["tarefas_criticas_atrasadas"])],
        ["Total de Tarefas", str(consolidado["total_tarefas"])],
    ]
    tabela_ind = Table(dados_indicadores, colWidths=[9 * cm, 7 * cm])
    tabela_ind.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E4053")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]
        )
    )
    elementos.append(tabela_ind)
    elementos.append(Spacer(1, 0.7 * cm))

    elementos.append(Paragraph("Curva S Consolidada do Portfólio", estilos["Heading2"]))
    if not curva_portfolio.empty:
        png_curva = _grafico_curva_s_portfolio_png(curva_portfolio)
        elementos.append(_imagem_proporcional(png_curva, largura_util / cm, 9))
    else:
        elementos.append(Paragraph("Não foi possível gerar a Curva S consolidada.", estilos["Normal"]))
    elementos.append(Spacer(1, 0.7 * cm))

    elementos.append(Paragraph("Comparativo entre Projetos", estilos["Heading2"]))
    colunas = list(tabela_comparativa.columns)

    estilo_cabecalho = ParagraphStyle("cabecalho_comp", parent=estilos["Normal"], fontSize=7, textColor=colors.white, leading=8.5)
    estilo_celula = ParagraphStyle("celula_comp", parent=estilos["Normal"], fontSize=7, leading=8.5)

    # Colunas com texto tipicamente mais longo (nome do projeto, período da linha de
    # base) recebem mais espaço; as demais (datas curtas e números) dividem o restante.
    # Todas as células usam Paragraph para quebrar linha em vez de estourar e sobrepor
    # a coluna vizinha.
    largura_larga = 3.2 * cm
    colunas_largas = {colunas[0], "Linha de Base Ativa", "Active Baseline"}
    n_largas = sum(1 for c in colunas if c in colunas_largas)
    n_estreitas = len(colunas) - n_largas
    largura_estreita = (largura_util - largura_larga * n_largas) / n_estreitas if n_estreitas else largura_util / len(colunas)
    larguras = [largura_larga if c in colunas_largas else largura_estreita for c in colunas]

    dados_tabela = [[Paragraph(str(c), estilo_cabecalho) for c in colunas]]
    for _, linha in tabela_comparativa.iterrows():
        dados_tabela.append([Paragraph(str(v) if v is not None else "N/D", estilo_celula) for v in linha])

    tabela_comp = Table(dados_tabela, colWidths=larguras, repeatRows=1)
    tabela_comp.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E4053")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elementos.append(tabela_comp)

    doc.build(elementos)
    return buffer.getvalue()


def _grafico_curva_s_png(curva_s: pd.DataFrame, unidade: str) -> bytes:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(16, 8))
    colunas_grafico = ["Linha de Base", "Realizado / Previsto"]
    for coluna in colunas_grafico:
        estilo = "--" if coluna == "Linha de Base" else "-"
        ax.plot(curva_s.index, curva_s[coluna], label=coluna, linewidth=2, linestyle=estilo)
    ax.set_title("Curva S - Linha de Base x Realizado/Previsto")
    ax.set_xlabel("Data")
    ax.set_ylabel("Custo acumulado" if unidade == "R$" else "Trabalho acumulado (horas)")
    ax.legend()
    ax.grid(True, alpha=0.3)
    fig.autofmt_xdate()

    saida = io.BytesIO()
    fig.savefig(saida, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    return saida.getvalue()


def gerar_pdf(
    projeto: Projeto,
    indicadores: Indicadores,
    percepcoes: list[str],
    curva_s: pd.DataFrame,
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    estilos = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph(f"Relatório de Andamento — {projeto.nome}", estilos["Title"]))
    elementos.append(Spacer(1, 0.5 * cm))

    rotulo_cpi = "CPI (índice de custo)" if indicadores.unidade == "R$" else "CPI (índice de eficiência)"
    dados_indicadores = [
        ["Indicador", "Valor"],
        ["SPI (índice de prazo)", f"{indicadores.spi:.2f}" if indicadores.spi is not None else "N/D"],
        [rotulo_cpi, f"{indicadores.cpi:.2f}" if indicadores.cpi is not None else "N/D"],
        ["% Concluído", f"{indicadores.percentual_concluido:.1f}%"],
        ["Atraso (dias)", str(indicadores.atraso_dias)],
        ["SV (variação de prazo, EV - PV)", formatar_valor(indicadores.variacao_prazo, indicadores.unidade)],
        ["CV (variação de custo, EV - AC)", formatar_valor(indicadores.variacao_custo, indicadores.unidade)],
        ["Tarefas Atrasadas", f"{indicadores.tarefas_atrasadas} de {indicadores.total_tarefas}"],
        ["Tarefas Críticas Atrasadas", str(indicadores.tarefas_criticas_atrasadas)],
    ]
    tabela = Table(dados_indicadores, colWidths=[9 * cm, 7 * cm])
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E4053")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]
        )
    )
    elementos.append(tabela)
    elementos.append(Spacer(1, 0.7 * cm))

    elementos.append(Paragraph("Principais Percepções", estilos["Heading2"]))
    for texto in percepcoes:
        elementos.append(Paragraph(f"• {texto}", estilos["Normal"]))
    elementos.append(Spacer(1, 0.7 * cm))

    elementos.append(Paragraph("Curva S - Linha de Base x Realizado/Previsto", estilos["Heading2"]))
    png = _grafico_curva_s_png(curva_s, indicadores.unidade)
    elementos.append(Image(io.BytesIO(png), width=17 * cm, height=8.5 * cm))

    doc.build(elementos)
    return buffer.getvalue()


def _cor_cartao(bom: bool | None):
    if bom is None:
        return colors.HexColor("#EAECEE")
    return colors.HexColor("#D4EFDF") if bom else colors.HexColor("#F8D7DA")


def gerar_pdf_status_reuniao(
    projeto: Projeto,
    indicadores: Indicadores,
    curva_s: pd.DataFrame,
    riscos: list[str],
    marcos_pendentes: list[Tarefa],
    tarefas_atrasadas_top: list[Tarefa],
    resultado_checklist: dict,
    periodo_linha_base_ativa_texto: str,
    data_status,
) -> bytes:
    """Relatório de uma única página no mesmo formato da aba 'Status de Reunião' da
    interface (números principais, mini Curva S, riscos, marcos e tarefas mais
    atrasadas) — pensado para impressão ou compartilhamento rápido em reunião."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=1 * cm, bottomMargin=1 * cm, leftMargin=1.2 * cm, rightMargin=1.2 * cm,
    )
    estilos = getSampleStyleSheet()
    largura_util = landscape(A4)[0] - 2.4 * cm
    elementos = []

    elementos.append(Paragraph(f"Status de Reunião — {projeto.nome}", estilos["Title"]))
    elementos.append(
        Paragraph(
            f"Data de status: {data_status:%d/%m/%Y} · Linha de base ativa: {periodo_linha_base_ativa_texto} · "
            f"Qualidade do cronograma: {resultado_checklist['classificacao']} "
            f"({resultado_checklist['percentual']:.0f}%)",
            estilos["Normal"],
        )
    )
    elementos.append(Spacer(1, 0.4 * cm))

    valores_kpi = [
        f"{indicadores.percentual_concluido:.1f}%",
        f"{indicadores.spi:.2f}" if indicadores.spi is not None else "N/D",
        f"{indicadores.cpi:.2f}" if indicadores.cpi is not None else "N/D",
        f"{indicadores.atraso_dias} dia(s)",
        str(indicadores.tarefas_criticas_atrasadas),
    ]
    rotulos_kpi = ["% Concluído", "SPI", "CPI", "Atraso", "Críticas Atrasadas"]
    bons_kpi = [
        None,
        (indicadores.spi >= 0.95) if indicadores.spi is not None else None,
        (indicadores.cpi >= 0.95) if indicadores.cpi is not None else None,
        indicadores.atraso_dias <= 0,
        indicadores.tarefas_criticas_atrasadas == 0,
    ]
    tabela_kpi = Table([valores_kpi, rotulos_kpi], colWidths=[largura_util / 5] * 5)
    estilo_kpi = [
        ("FONTSIZE", (0, 0), (-1, 0), 20),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.white),
    ]
    for col, bom in enumerate(bons_kpi):
        estilo_kpi.append(("BACKGROUND", (col, 0), (col, 1), _cor_cartao(bom)))
    tabela_kpi.setStyle(TableStyle(estilo_kpi))
    elementos.append(tabela_kpi)
    elementos.append(Spacer(1, 0.5 * cm))

    png_curva = _grafico_curva_s_png(curva_s, indicadores.unidade)
    imagem_curva = _imagem_proporcional(png_curva, largura_util / cm * 0.55, 9)

    estilo_lista = ParagraphStyle("lista_status", parent=estilos["Normal"], fontSize=8.5, leading=11)
    bloco_direita = [Paragraph("Principais riscos", estilos["Heading3"])]
    if riscos:
        for texto in riscos:
            bloco_direita.append(Paragraph(f"• {texto}", estilo_lista))
    else:
        bloco_direita.append(Paragraph("Nenhum risco identificado no momento.", estilo_lista))
    bloco_direita.append(Spacer(1, 0.3 * cm))
    bloco_direita.append(Paragraph("Próximos marcos", estilos["Heading3"]))
    if marcos_pendentes:
        for marco in marcos_pendentes:
            data_marco = f"{marco.termino:%d/%m/%Y}" if marco.termino else "N/D"
            bloco_direita.append(Paragraph(f"• {marco.nome} — {data_marco}", estilo_lista))
    else:
        bloco_direita.append(Paragraph("Nenhum marco pendente.", estilo_lista))

    tabela_duas_colunas = Table(
        [[imagem_curva, bloco_direita]],
        colWidths=[largura_util * 0.55, largura_util * 0.45],
    )
    tabela_duas_colunas.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    elementos.append(tabela_duas_colunas)
    elementos.append(Spacer(1, 0.5 * cm))

    elementos.append(
        Paragraph(f"Tarefas mais atrasadas ({indicadores.tarefas_atrasadas})", estilos["Heading3"])
    )
    if tarefas_atrasadas_top:
        dados_tabela = [["Tarefa", "Crítica", "Atraso (dias)", "% Concluído"]]
        for t in tarefas_atrasadas_top:
            dados_tabela.append(
                [t.nome, "Sim" if t.critica else "Não", str(dias_atraso_tarefa(t)), f"{t.percentual_concluido:.0f}%"]
            )
        tabela_atrasadas = Table(
            dados_tabela,
            colWidths=[largura_util * 0.55, largura_util * 0.15, largura_util * 0.15, largura_util * 0.15],
            repeatRows=1,
        )
        tabela_atrasadas.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E4053")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ]
            )
        )
        elementos.append(tabela_atrasadas)
    else:
        elementos.append(Paragraph("Nenhuma tarefa está atrasada em relação à linha de base.", estilos["Normal"]))

    doc.build(elementos)
    return buffer.getvalue()


def _grafico_gantt_png(tarefas: list[Tarefa], data_status, max_tarefas: int = 40):
    """Gera uma imagem PNG de um Gantt simples (matplotlib) para uso em PDF.

    Retorna (bytes_png, tarefas_exibidas, total_original). Se houver mais tarefas do que
    'max_tarefas', prioriza as críticas/atrasadas e sinaliza o corte via 'limitado'.
    """
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.dates as mdates
    import matplotlib.pyplot as plt

    total_original = len(tarefas)
    tarefas_ordenadas = sorted(tarefas, key=lambda t: (t.inicio, t.id))
    if len(tarefas_ordenadas) > max_tarefas:
        tarefas_ordenadas = sorted(
            tarefas_ordenadas,
            key=lambda t: (not (t.critica and t.atrasada), not t.atrasada, not t.critica, t.inicio),
        )[:max_tarefas]
        tarefas_ordenadas.sort(key=lambda t: (t.inicio, t.id))

    n = len(tarefas_ordenadas)
    fig, ax = plt.subplots(figsize=(16, max(4, 0.32 * n)))

    for i, t in enumerate(tarefas_ordenadas):
        y = n - 1 - i
        cor = CORES_STATUS[status_tarefa(t)]
        if t.marco:
            ax.plot(t.termino or t.inicio, y, marker="D", color=cor, markersize=8, markeredgecolor="white", markeredgewidth=0.6)
        else:
            largura = max((t.termino - t.inicio).days, 1)
            ax.barh(y, largura, left=t.inicio, color=cor, edgecolor="white", linewidth=0.3, height=0.6)

    ax.set_yticks(range(n))
    ax.set_yticklabels([f"{t.id} - {t.nome}" for t in reversed(tarefas_ordenadas)], fontsize=7)
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b/%Y"))
    if data_status:
        ax.axvline(data_status, color="gray", linestyle=":", linewidth=1)
    ax.set_xlabel("Data")
    ax.grid(True, axis="x", alpha=0.3)
    fig.autofmt_xdate()

    handles = [plt.Line2D([0], [0], color=cor, lw=6, label=nome) for nome, cor in CORES_STATUS.items()]
    ax.legend(handles=handles, loc="upper left", bbox_to_anchor=(1.005, 1), fontsize=8)
    fig.tight_layout()

    saida = io.BytesIO()
    fig.savefig(saida, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    return saida.getvalue(), n, total_original


def _imagem_proporcional(png_bytes: bytes, largura_cm: float, altura_max_cm: float) -> Image:
    from PIL import Image as PILImage

    buffer_img = io.BytesIO(png_bytes)
    with PILImage.open(buffer_img) as im:
        proporcao = im.height / im.width
    buffer_img.seek(0)
    altura_cm = min(largura_cm * proporcao, altura_max_cm)
    return Image(buffer_img, width=largura_cm * cm, height=altura_cm * cm)


def gerar_pdf_executivo(
    projeto: Projeto,
    indicadores: Indicadores,
    curva_periodo: pd.DataFrame,
    tarefas_periodo: list[Tarefa],
    periodo_inicio,
    periodo_fim,
    data_status,
) -> bytes:
    """Relatório executivo focado em um período: indicadores, Curva S, Gantt e atividades."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    estilos = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph(f"Relatório Executivo — {projeto.nome}", estilos["Title"]))
    elementos.append(Paragraph(f"Período: {periodo_inicio:%d/%m/%Y} a {periodo_fim:%d/%m/%Y}", estilos["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))

    rotulo_cpi = "CPI (índice de custo)" if indicadores.unidade == "R$" else "CPI (índice de eficiência)"
    dados_indicadores = [
        ["Indicador", "Valor"],
        ["SPI (índice de prazo)", f"{indicadores.spi:.2f}" if indicadores.spi is not None else "N/D"],
        [rotulo_cpi, f"{indicadores.cpi:.2f}" if indicadores.cpi is not None else "N/D"],
        ["% Concluído (geral do projeto)", f"{indicadores.percentual_concluido:.1f}%"],
        ["Atraso (dias)", str(indicadores.atraso_dias)],
        ["Tarefas no período", str(len(tarefas_periodo))],
    ]
    tabela_ind = Table(dados_indicadores, colWidths=[9 * cm, 7 * cm])
    tabela_ind.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E4053")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]
        )
    )
    elementos.append(tabela_ind)
    elementos.append(Spacer(1, 0.7 * cm))

    elementos.append(Paragraph("Curva S no período", estilos["Heading2"]))
    if not curva_periodo.empty:
        png_curva = _grafico_curva_s_png(curva_periodo, indicadores.unidade)
        elementos.append(_imagem_proporcional(png_curva, 17, 9))
    else:
        elementos.append(Paragraph("Sem dados de curva no período selecionado.", estilos["Normal"]))
    elementos.append(Spacer(1, 0.7 * cm))

    elementos.append(Paragraph("Gráfico de Gantt no período", estilos["Heading2"]))
    if tarefas_periodo:
        png_gantt, exibidas, total_original = _grafico_gantt_png(tarefas_periodo, data_status)
        elementos.append(_imagem_proporcional(png_gantt, 17, 20))
        if exibidas < total_original:
            elementos.append(
                Paragraph(
                    f"Exibindo {exibidas} de {total_original} tarefas (priorizadas por criticidade/atraso) "
                    "para manter o gráfico legível.",
                    estilos["Italic"],
                )
            )
    else:
        elementos.append(Paragraph("Nenhuma tarefa com datas no período selecionado.", estilos["Normal"]))
    elementos.append(Spacer(1, 0.7 * cm))

    elementos.append(Paragraph(f"Atividades no período ({len(tarefas_periodo)})", estilos["Heading2"]))
    if tarefas_periodo:
        dados_tabela = [["ID", "Tarefa", "Início", "Término", "% Concl.", "Status"]]
        for t in sorted(tarefas_periodo, key=lambda t: (t.inicio, t.id)):
            nome = t.nome.strip()
            if len(nome) > 55:
                nome = nome[:54] + "…"
            dados_tabela.append(
                [
                    str(t.id),
                    nome,
                    t.inicio.strftime("%d/%m/%Y") if t.inicio else "",
                    t.termino.strftime("%d/%m/%Y") if t.termino else "",
                    f"{t.percentual_concluido:.0f}%",
                    status_tarefa(t),
                ]
            )
        tabela_atividades = Table(
            dados_tabela, colWidths=[1.2 * cm, 6.5 * cm, 2.3 * cm, 2.3 * cm, 1.7 * cm, 3 * cm], repeatRows=1
        )
        tabela_atividades.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E4053")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ]
            )
        )
        elementos.append(tabela_atividades)
    else:
        elementos.append(Paragraph("Nenhuma atividade encontrada no período selecionado.", estilos["Normal"]))

    doc.build(elementos)
    return buffer.getvalue()
